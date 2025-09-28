import streamlit as st
import pdfplumber
import json
import pandas as pd
import re
import os
import tempfile
from pathlib import Path
import numpy as np
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# Set page configuration
st.set_page_config(
    page_title="PDF to Excel Converter - SLIK Report",
    page_icon="📊",
    layout="wide"
)

# Custom CSS untuk styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .sub-header {
        font-size: 1.5rem;
        color: #2e86ab;
        margin-top: 1.5rem;
        margin-bottom: 1rem;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 5px;
        padding: 15px;
        margin: 10px 0;
    }
    .info-box {
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        border-radius: 5px;
        padding: 15px;
        margin: 10px 0;
    }
</style>
""", unsafe_allow_html=True)

# Judul aplikasi
st.markdown('<div class="main-header">🔄 PDF to Excel Converter - SLIK Report</div>', unsafe_allow_html=True)

# Sidebar untuk upload file
st.sidebar.title("📁 Upload PDF Files")
uploaded_files = st.sidebar.file_uploader(
    "Pilih file PDF SLIK Report", 
    type=["pdf"], 
    accept_multiple_files=True
)

# Fungsi-fungsi utama (diperbaiki sesuai kode ipynb)
def pdf_to_json(pdf_file, json_file_path):
    """Mengkonversi file PDF ke JSON"""
    text_data = []  # LIST SIMPAN TEKS DARI SETIAP PAGE

    with pdfplumber.open(pdf_file) as pdf:  
        for page in pdf.pages:  
            text = page.extract_text()  
            if text: 
                text_data.append(text) 

    # SIMPAN TEKS DALAM FORMAT JSON
    with open(json_file_path, 'w', encoding='utf-8') as output_file:  
        json.dump(text_data, output_file, ensure_ascii=False, indent=4)
    
    return text_data

def read_json_files(json_files):
    """Membaca semua file JSON dan menggabungkannya"""
    dataframes = []
    for json_file in json_files:
        try:
            df = pd.read_json(json_file)  
            dataframes.append(df)
        except Exception as e:
            st.error(f"Gagal membaca file {json_file}: {str(e)}")
    
    if not dataframes:
        return None
    
    combined_df = pd.concat(dataframes, ignore_index=True)
    return combined_df

def process_kredit_data(combined_data):
    """Memproses data kredit dari data gabungan"""
    # Pilih Baris == 'Informasi Debitur' atau 'Jenis Kredit/Pembiayaan'
    kredit = combined_data[combined_data[0].str.contains('Sistem Layanan Informasi Keuangan|Jenis Kredit/Pembiayaan', na=False)]
    kredit = pd.DataFrame(kredit)
    lines = kredit[0].apply(lambda x: x.split('\n')).explode().tolist()
    data_list = []
    
    # Dictionary untuk Kolom
    data_dict = {
        "BANK": None,
        "Baki Debet": None,
        "No Rekening": None,
        "Kualitas": None,
        "Jenis Kredit/Pembiayaan": None,
        "Akad Kredit/Pembiayaan": None,
        "Frekuensi Perpanjangan Kredit/": None,
        "No Akad Awal": None,
        "Tanggal Akad Awal": None,
        "No Akad Akhir": None,
        "Tanggal Akad Akhir": None,
        "Tanggal Awal Kredit": None,
        "Tanggal Mulai": None,
        "Tanggal Jatuh Tempo": None,
        "Kategori Debitur": None,
        "Jenis Penggunaan": None,
        "Sektor Ekonomi": None,
        "Kredit Program Pemerintah": None,
        "Kab/Kota Lokasi Proyek": None,
        "Valuta": None,
        "Suku Bunga/Imbalan": None,
        "Jenis Suku Bunga/Imbalan": None,
        "Keterangan": None,
        "Nama Debitur": None,
        "Nama Group": None
    }
    
    # Variabel untuk Save Nama Deb & Group
    nama_debitur = None
    nama_group = None

    for i, line in enumerate(lines):
        if line.strip() == "Nomor Laporan":  # Hanya baris yang persis "Nomor Laporan"
            if i + 2 >= 0:  # Pastikan indeks tidak negatif
                group_line = lines[i + 2].strip()
                nama_group = " ".join(group_line.split()[:3])
                data_dict["Nama Group"] = nama_group
                
        # Baris == "Penyajian informasi debitur pada Sistem Layanan Informasi"
        if "Penyajian informasi debitur pada Sistem Layanan Informasi" in line:
            # Ambil Baris Ke-4 Setelah Teks Tersebut
            if i + 3 < len(lines):
                nama_line = lines[i + 3].strip()  
                # Ambil 5 Kata Pertama
                nama_debitur = " ".join(nama_line.split()[:5])
                data_dict["Nama Debitur"] = nama_debitur

        # NAMA PELAPOR/BANK
        if " - " in line and "Rp" in line:
            if i > 0 and ("Pelapor Cabang" in lines[i-1]):
                bank_part = line.split("Rp")[0].strip()
                data_dict["BANK"] = bank_part

        # BAKI DEBET
            os_part = line.split("Rp")[1].strip()
            os_value = os_part.split(" ")[0]
            data_dict["Baki Debet"] = os_value

        # NO REKENING & KUALITAS
        elif "No Rekening" in line and "Kualitas" in line:
            nokredit_parts = line.split("No Rekening")[1].strip("Kualitas")[0].strip()
            data_dict["No Rekening"] = nokredit_parts
            kualitas_parts = line.split("Kualitas")[1].strip()
            data_dict["Kualitas"] = kualitas_parts
        
        # SIFAT KREDIT/PEMBIAYAAN & JUMLAH HARI TUNGGAKAN
        elif "Sifat Kredit/Pembiayaan" in line and "Jumlah Hari Tunggakan":
            sifatkre_parts = line.split("Sifat Kredit/Pembiayaan")[1].strip()
            data_dict["Sifat Kredit/Pembiayaan"] = sifatkre_parts

        # JENIS KREDIT/PEMBIAYAAN & NILAI PROYEK
        elif "Jenis Kredit/Pembiayaan" in line and "Nilai Proyek":
            jeniskre_parts = line.split("Jenis Kredit/Pembiayaan")[1].strip()
            data_dict["Jenis Kredit/Pembiayaan"] = jeniskre_parts

        # AKAD KREDIT/PEMBIAYAAN & PLAFON AWAL
        elif "Akad Kredit/Pembiayaan" in line and "Plafon Awal":
            akadkre_parts = line.split("Akad Kredit/Pembiayaan")[1].strip()
            data_dict["Akad Kredit/Pembiayaan"] = akadkre_parts

        # FREKUENSI PERPANJANGAN KREDIT/ & PLAFON
        elif "Frekuensi Perpanjangan Kredit/" in line and "Plafon":
            frekre_parts = line.split("Frekuensi Perpanjangan Kredit/")[1].strip()
            data_dict["Frekuensi Perpanjangan Kredit/"] = frekre_parts

        # NO AKAD AWAL & REALISASI/PENCAIRAN BULAN BERJALAN
        elif "No Akad Awal" in line and "Realisasi/Pencairan Bulan Berjalan":
            noakadawal_parts = line.split("No Akad Awal")[1].strip()
            data_dict["No Akad Awal"] = noakadawal_parts

        # TANGGAL AKAD AWAL & NILAI DALAM MATA UANG ASAL
        elif "Tanggal Akad Awal" in line and "Nilai dalam Mata Uang Asal":
            tglakadawal_parts = line.split("Tanggal Akad Awal")[1].strip()
            data_dict["Tanggal Akad Awal"] = tglakadawal_parts

        # NO AKAD AKHIR & SEBAB MACET
        elif "No Akad Akhir" in line and "Sebab Macet":
            noakadakhir_parts = line.split("No Akad Akhir")[1].strip()
            data_dict["No Akad Akhir"] = noakadakhir_parts

        # TANGGAL AKAD AKHIR & TANGGAL MACET
        elif "Tanggal Akad Akhir" in line and "Tanggal Macet":
            tglakadakhir_parts = line.split("Tanggal Akad Akhir")[1].strip()
            data_dict["Tanggal Akad Akhir"] = tglakadakhir_parts

        # TANGGAL AWAL KREDIT & TUNGGAKAN POKOK
        elif "Tanggal Awal Kredit" in line and "Tunggakan Pokok":
            tglawalkre_parts = line.split("Tanggal Awal Kredit")[1].strip()
            data_dict["Tanggal Awal Kredit"] = tglawalkre_parts

        # TANGGAL MULAI & TUNGGAKAN BUNGA
        elif "Tanggal Mulai" in line and "Tunggakan Bunga":
            tglmulai_parts = line.split("Tanggal Mulai")[1].strip()
            data_dict["Tanggal Mulai"] = tglmulai_parts

        # TANGGAL JATUH TEMPO & FREKUENSI TUNGGAKAN
        elif "Tanggal Jatuh Tempo" in line and "Frekuensi Tunggakan":
            tgljatem_parts = line.split("Tanggal Jatuh Tempo")[1].strip()
            data_dict["Tanggal Jatuh Tempo"] = tgljatem_parts

        # KATEGORI DEBITUR & DENDA
        elif "Kategori Debitur" in line and "Denda":
            katdeb_parts = line.split("Kategori Debitur")[1].strip()
            data_dict["Kategori Debitur"] = katdeb_parts

        # JENIS PENGGUNAAN & FREKUENSI RESTRUKTURISASI
        elif "Jenis Penggunaan" in line and "Frekuensi Restrukturisasi":
            jenispeng_parts = line.split("Jenis Penggunaan")[1].strip()
            data_dict["Jenis Penggunaan"] = jenispeng_parts

        # SEKTOR EKONOMI & TANGGAL RESTRUKTURISASI AKHIR
        elif "Sektor Ekonomi" in line and "Tanggal Restrukturisasi Akhir":
            sekon_parts = line.split("Sektor Ekonomi")[1].strip()
            data_dict["Sektor Ekonomi"] = sekon_parts

        # KREDIT PROGRAM PEMERINTAH & CARA RESTRUKTURISASI
        elif "Kredit Program Pemerintah" in line and "Cara Restrukturisasi":
            kreditprogram_parts = line.split("Kredit Program Pemerintah")[1].strip()
            data_dict["Kredit Program Pemerintah"] = kreditprogram_parts

        # KAB/KOTA LOKASI PROYEK & KONDISI
        elif "Kab/Kota Lokasi Proyek" in line and "Kondisi":
            lokasi_parts = line.split("Kab/Kota Lokasi Proyek")[1].strip()
            data_dict["Kab/Kota Lokasi Proyek"] = lokasi_parts

        # VALUTA & TANGGAL KONDISI
        elif "Valuta" in line and "Tanggal Kondisi":
            valuta_parts = line.split("Valuta")[1].strip()
            data_dict["Valuta"] = valuta_parts

        # SUKU BUNGA/IMBALAN & JENIS BUNGA/IMBALAN
        elif "Suku Bunga/Imbalan" in line and "Jenis Suku Bunga/Imbalan" in line:  
            parts = line.split(" ")  
            data_dict["Suku Bunga/Imbalan"] = parts[2]
            jenisbunga_parts = line.split("Jenis Suku Bunga/Imbalan")[1].strip()
            data_dict["Jenis Suku Bunga/Imbalan"] = jenisbunga_parts

        # KETERANGAN
        elif "Keterangan" in line:
            # Cek apakah baris sebelumnya mengandung Bank Beneficiary
            if i > 0 and ("Jenis Suku Bunga/Imbalan" in lines[i-1]):
                keterangan_parts = line.split("Keterangan")[1].strip()
                data_dict["Keterangan"] = keterangan_parts
                
                # MENYIMPAN DICTIONARY KEDALAM LIST
                data_list.append(data_dict.copy())

    kredit = pd.DataFrame(data_list)
    
    if not kredit.empty:
        kredit = kredit.dropna(subset=['Jenis Kredit/Pembiayaan'])

        # Split kolom-kolom yang perlu dipisah
        if 'Sifat Kredit/Pembiayaan' in kredit.columns:
            kredit[['Sifat Kredit/Pembiayaan', 'Jumlah Hari Tunggakan']] = kredit['Sifat Kredit/Pembiayaan'].apply(
                lambda x: pd.Series(x.split('Jumlah Hari Tunggakan', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Jenis Kredit/Pembiayaan' in kredit.columns:
            kredit[['Jenis Kredit/Pembiayaan', 'Nilai Proyek']] = kredit['Jenis Kredit/Pembiayaan'].apply(
                lambda x: pd.Series(x.split('Nilai Proyek', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Akad Kredit/Pembiayaan' in kredit.columns:
            kredit[['Akad Kredit/Pembiayaan', 'Plafon Awal']] = kredit['Akad Kredit/Pembiayaan'].apply(
                lambda x: pd.Series(x.split('Plafon Awal', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Frekuensi Perpanjangan Kredit/' in kredit.columns:
            kredit[['Frekuensi Perpanjangan Kredit/', 'Plafon']] = kredit['Frekuensi Perpanjangan Kredit/'].apply(
                lambda x: pd.Series(x.split('Plafon', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'No Akad Awal' in kredit.columns:
            kredit[['No Akad Awal', 'Realisasi/Pencairan Bulan Berjalan']] = kredit['No Akad Awal'].apply(
                lambda x: pd.Series(x.split('Realisasi/Pencairan Bulan Berjalan', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Tanggal Akad Awal' in kredit.columns:
            kredit[['Tanggal Akad Awal', 'Nilai dalam Mata Uang Asal']] = kredit['Tanggal Akad Awal'].apply(
                lambda x: pd.Series(x.split('Nilai dalam Mata Uang Asal', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'No Akad Akhir' in kredit.columns:
            kredit[['No Akad Akhir', 'Sebab Macet']] = kredit['No Akad Akhir'].apply(
                lambda x: pd.Series(x.split('Sebab Macet', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Tanggal Akad Akhir' in kredit.columns:
            kredit[['Tanggal Akad Akhir', 'Tanggal Macet']] = kredit['Tanggal Akad Akhir'].apply(
                lambda x: pd.Series(x.split('Tanggal Macet', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Tanggal Awal Kredit' in kredit.columns:
            kredit[['Tanggal Awal Kredit', 'Tunggakan Pokok']] = kredit['Tanggal Awal Kredit'].apply(
                lambda x: pd.Series(x.split('Tunggakan Pokok', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Tanggal Mulai' in kredit.columns:
            kredit[['Tanggal Mulai', 'Tunggakan Bunga']] = kredit['Tanggal Mulai'].apply(
                lambda x: pd.Series(x.split('Tunggakan Bunga', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Tanggal Jatuh Tempo' in kredit.columns:
            kredit[['Tanggal Jatuh Tempo', 'Frekuensi Tunggakan']] = kredit['Tanggal Jatuh Tempo'].apply(
                lambda x: pd.Series(x.split('Frekuensi Tunggakan', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Kategori Debitur' in kredit.columns:
            kredit[['Kategori Debitur', 'Denda']] = kredit['Kategori Debitur'].apply(
                lambda x: pd.Series(x.split('Denda', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Jenis Penggunaan' in kredit.columns:
            kredit[['Jenis Penggunaan', 'Frekuensi Restrukturisasi']] = kredit['Jenis Penggunaan'].apply(
                lambda x: pd.Series(x.split('Frekuensi Restrukturisasi', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Sektor Ekonomi' in kredit.columns:
            kredit[['Sektor Ekonomi', 'Tanggal Restrukturisasi Akhir']] = kredit['Sektor Ekonomi'].apply(
                lambda x: pd.Series(x.split('Tanggal Restrukturisasi Akhir', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Kredit Program Pemerintah' in kredit.columns:
            kredit[['Kredit Program Pemerintah', 'Cara Restrukturisasi']] = kredit['Kredit Program Pemerintah'].apply(
                lambda x: pd.Series(x.split('Cara Restrukturisasi', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Kab/Kota Lokasi Proyek' in kredit.columns:
            kredit[['Kab/Kota Lokasi Proyek', 'Kondisi']] = kredit['Kab/Kota Lokasi Proyek'].apply(
                lambda x: pd.Series(x.split('Kondisi', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        if 'Valuta' in kredit.columns:
            kredit[['Valuta', 'Tanggal Kondisi']] = kredit['Valuta'].apply(
                lambda x: pd.Series(x.split('Tanggal Kondisi', 1)) if pd.notna(x) else pd.Series([None, None])
            )

        # Filter data
        if 'Keterangan' in kredit.columns:
            kredit = kredit[~kredit['Keterangan'].str.contains('Tgl Penilaian Penilai Independen|Garansi|L/C', na=False, case=False)]

        # Cleaning persentase
        def clean_percentage(value):
            if pd.isnull(value):
                return value 
            value = str(value).replace('%', '').strip()
            try:
                return float(value)
            except ValueError:
                return None

        if 'Suku Bunga/Imbalan' in kredit.columns:
            kredit['Suku Bunga/Imbalan'] = kredit['Suku Bunga/Imbalan'].apply(clean_percentage)

        # Standardisasi nilai
        if 'Jenis Kredit/Pembiayaan' in kredit.columns:
            kredit['Jenis Kredit/Pembiayaan'] = kredit['Jenis Kredit/Pembiayaan'].str.strip().replace(
                {
                    'Kredit atau Pembiayaan untuk': 'Kredit atau Pembiayaan untuk Pembayaran Bersama (Sindikasi)',
                    'Kartu Kredit atau Kartu Pembiayaan': 'Kartu Kredit atau Kartu Pembiayaan Syariah',
                    'Kredit atau Pembiayaan kepada Pihak': 'Kredit atau Pembiayaan kepada Pihak Ketiga Melalui Lembaga Lain Secara Channeling',
                    'Kredit atau Pembiayaan kepada Non-UMKM': 'Kredit atau Pembiayaan kepada Non-UMKM melalui Lembaga Lain Secara Executing',
                    'Kredit atau Pembiayaan kepada UMKM': 'Kredit atau Pembiayaan kepada UMKM Melalui Lembaga Lain Secara Executing',
                    'Kredit/ Pembiayaan Kepada Non-UMKM': 'Kredit atau Pembiayaan kepada Non-UMKM melalui Lembaga Lain Secara Executing',
                    'Kredit/Pembiayaan Dalam Rangka': 'Kredit atau Pembiayaan Dalam Rangka Pembiayaan Bersama (Sindikasi)',
                },
                regex=False)

        if 'Kategori Debitur' in kredit.columns:
            kredit['Kategori Debitur'] = kredit['Kategori Debitur'].str.strip().replace(
                {'Bukan Debitur Usaha Mikro, Kecil, dan': 'Bukan Debitur Usaha Mikro, Kecil, dan Menengah'},
                regex=False)

        if 'Sektor Ekonomi' in kredit.columns:
            kredit['Sektor Ekonomi'] = kredit['Sektor Ekonomi'].str.strip().replace({
                'Industri Rokok dan Produk Tembakau': 'Industri Rokok dan Produk Tembakau Lainnya',
                'Industri Penggilingan Beras dan Jagung': 'Industri Penggilingan Beras dan Jagung dan Industri Tepung Beras dan Jagung',
                'Industri Penggilingan Padi dan': 'Industri Penggilingan Padi dan Penyosohan Beras',
                'Perdagangan Besar Mesin-mesin, Suku': 'Perdagangan Besar Mesin-mesin, Suku Cadang dan Perlengkapannya',
                'Perdagangan Eceran Mesin-mesin': 'Perdagangan Eceran Mesin-mesin (Kecuali Mobil dan Sepeda Motor) dan Suku Cadang, termasuk Alat-alat Tranportasi',
                'Perdagangan Besar Mesin, Peralatan': 'Perdagangan Besar Mesin, Peralatan dan Perlengkapannya',
                'Perdagangan Impor Suku Cadang': 'Perdagangan Impor Suku Cadang Mesin-mesin, Suku Cadang dan Perlengkapan Lain',
                'Rumah Tangga Untuk Pemilikan Mobil': 'Rumah Tangga Untuk Pemilikian Mobil Roda Empat',
            }, regex=False)

        if 'Kredit Program Pemerintah' in kredit.columns:
            kredit['Kredit Program Pemerintah'] = kredit['Kredit Program Pemerintah'].str.strip().replace({
                'Kredit yang bukan merupakan kredit/': 'Kredit yang bukan merupakan kredit/pembiayaan dalam rangka program pemerintah'
            }, regex=False)

        # Tambahkan kategori dan rename kolom
        kredit['Kategori'] = 'Kredit/Pembiayaan'
        
        finalkredit = kredit.rename(columns={
            "Baki Debet": "Baki Debet/Nominal",
            "No Rekening": "No Rek/LC/Surat",
            "Jenis Kredit/Pembiayaan": "Jenis Kredit/LC/Garansi/Surat/Fasilitas",
            "Tanggal Mulai": "Tanggal Mulai/Terbit",
            "Tanggal Macet": "Tanggal Macet/Wanprestasi",
            "Realisasi/Pencairan Bulan Berjalan": "Nilai Perolehan/Jaminan/Realisasi",
            "Jenis Penggunaan": "Tujuan/Jenis Penggunaan",
            "Nilai Proyek": "Nilai Pasar/Proyek",
            "Nilai dalam Mata Uang Asal": "Nilai Dalam Mata Uang Asal",
        })
        
        return finalkredit
    else:
        return pd.DataFrame()

def process_lc_data(combined_data):
    """Memproses data LC Irrecovable dari data gabungan"""
    # Pilih Baris == 'Informasi Debitur' atau 'Jenis L/C'
    lc = combined_data[combined_data[0].str.contains('Sistem Layanan Informasi Keuangan|Jenis L/C', na=False)]
    lc = pd.DataFrame(lc)
    lines = lc[0].apply(lambda x: x.split('\n')).explode().tolist()
    data_list = []
    data_dict = {
        "BANK": None,
        "Baki Debet": None,
        "No L/C": None,
        "Kualitas": None,
        "Jenis L/C": None,
        "Tanggal Keluar": None,
        "Tanggal Jatuh Tempo": None,
        "No Akad Awal": None,
        "Tanggal Akad Awal": None,
        "No Akad Akhir": None,
        "Tanggal Akad Akhir": None,
        "Bank Beneficiary": None,
        "Keterangan": None,
        "Nama Debitur": None,
        "Nama Group": None
    } 

    # Variabel untuk Save Nama Deb & Group
    nama_debitur = None
    nama_group = None

    for i, line in enumerate(lines):
        if line.strip() == "Nomor Laporan":  # Hanya baris yang persis "Nomor Laporan"
            if i + 2 >= 0:  # Pastikan indeks tidak negatif
                group_line = lines[i + 2].strip()
                nama_group = " ".join(group_line.split()[:3])
                data_dict["Nama Group"] = nama_group
                
        # Baris == "Penyajian informasi debitur pada Sistem Layanan Informasi"
        if "Penyajian informasi debitur pada Sistem Layanan Informasi" in line:
            # Ambil Baris Ke-4 Setelah Teks Tersebut
            if i + 3 < len(lines):
                nama_line = lines[i + 3].strip()  
                # Ambil 5 Kata Pertama
                nama_debitur = " ".join(nama_line.split()[:5])
                data_dict["Nama Debitur"] = nama_debitur

        # NAMA PELAPOR/BANK
        if " - " in line and "Rp" in line:
            if i > 0 and ("Pelapor Cabang" in lines[i-1]):
                bank_part = line.split("Rp")[0].strip()
                data_dict["BANK"] = bank_part

        # BAKI DEBET
            os_part = line.split("Rp")[1].strip()
            os_value = os_part.split(" ")[0]
            data_dict["Baki Debet"] = os_value

        # NO L/C & KUALITAS
        elif "No L/C" in line and "Kualitas" in line:
            nolc_parts = line.split("No L/C")[1].strip("Kualitas")[0].strip()
            data_dict["No L/C"] = nolc_parts
            kualitas_parts = line.split("Kualitas")[1].strip()
            data_dict["Kualitas"] = kualitas_parts
        
        # JENIS L/C
        elif "Jenis L/C" in line and "Valuta":
            jenislc_parts = line.split("Jenis L/C")[1].strip()
            data_dict["Jenis L/C"] = jenislc_parts

        # TANGGAL KELUAR
        elif "Tanggal Keluar" in line and "Plafon":
            tglkeluar_parts = line.split("Tanggal Keluar")[1].strip()
            data_dict["Tanggal Keluar"] = tglkeluar_parts

        # TANGGAL JATUH TEMPO
        elif "Tanggal Jatuh Tempo" in line and "Tujuan L/C":
            tgljatem_parts = line.split("Tanggal Jatuh Tempo")[1].strip()
            data_dict["Tanggal Jatuh Tempo"] = tgljatem_parts

        # NO AKAD AWAL
        elif "No Akad Awal" in line and "Setoran Jaminan" in line:
            noakadawal_parts = line.split("No Akad Awal")[1].strip()
            data_dict["No Akad Awal"] = noakadawal_parts

        # TANGGAL AKAD AWAL
        elif "Tanggal Akad Awal" in line and "Tanggal Wan Prestasi":
            tglakadawal_parts = line.split("Tanggal Akad Awal")[1].strip()
            data_dict["Tanggal Akad Awal"] = tglakadawal_parts

        # NO AKAD AKHIR
        elif "No Akad Akhir" in line and "Kondisi":
            noakadakhir_parts = line.split("No Akad Akhir")[1].strip()
            data_dict["No Akad Akhir"] = noakadakhir_parts

        # TANGGAL AKAD AKHIR
        elif "Tanggal Akad Akhir" in line and "Tanggal Kondisi":
            tglakadakhir_parts = line.split("Tanggal Akad Akhir")[1].strip()
            data_dict["Tanggal Akad Akhir"] = tglakadakhir_parts
        
        # BANK BENEFICIARY
        elif "Bank Beneficiary" in line:
            bankben_parts = line.split("Bank Beneficiary")[1].strip()
            data_dict["Bank Beneficiary"] = bankben_parts

        # KETERANGAN
        elif "Keterangan" in line:
            # Cek apakah baris sebelumnya mengandung Bank Beneficiary
            if i > 0 and ("Bank Beneficiary" in lines[i-1]):
                keterangan_parts = line.split("Keterangan")[1].strip()
                data_dict["Keterangan"] = keterangan_parts
                
                # MENYIMPAN DICTIONARY KEDALAM LIST
                data_list.append(data_dict.copy())

    lc = pd.DataFrame(data_list)

    if 'Jenis L/C' in lc.columns:
        lc = lc.dropna(subset=['Jenis L/C'])
    else:
        lc['Jenis L/C'] = pd.NA

    # KOLOM JENIS L/C & VALUTA
    if not lc.empty and 'Jenis L/C' in lc.columns:
        lc[['Jenis L/C', 'Valuta']] = lc['Jenis L/C'].apply(
            lambda x: pd.Series(x.split('Valuta', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not lc.empty and 'Tanggal Keluar' in lc.columns:
        lc[['Tanggal Keluar', 'Plafon']] = lc['Tanggal Keluar'].apply(
            lambda x: pd.Series(x.split('Plafon', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not lc.empty and 'Tanggal Jatuh Tempo' in lc.columns:
        lc[['Tanggal Jatuh Tempo', 'Tujuan L/C']] = lc['Tanggal Jatuh Tempo'].apply(
            lambda x: pd.Series(x.split('Tujuan L/C', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not lc.empty and 'No Akad Awal' in lc.columns:
        lc[['No Akad Awal', 'Setoran Jaminan']] = lc['No Akad Awal'].apply(
            lambda x: pd.Series(x.split('Setoran Jaminan', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not lc.empty and 'Tanggal Akad Awal' in lc.columns:
        lc[['Tanggal Akad Awal', 'Tanggal Wan Prestasi']] = lc['Tanggal Akad Awal'].apply(
            lambda x: pd.Series(x.split('Tanggal Wan Prestasi', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not lc.empty and 'No Akad Akhir' in lc.columns:
        lc[['No Akad Akhir', 'Kondisi']] = lc['No Akad Akhir'].apply(
            lambda x: pd.Series(x.split('Kondisi', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not lc.empty and 'Tanggal Akad Akhir' in lc.columns:
        lc[['Tanggal Akad Akhir', 'Tanggal Kondisi']] = lc['Tanggal Akad Akhir'].apply(
            lambda x: pd.Series(x.split('Tanggal Kondisi', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not lc.empty and 'Keterangan' in lc.columns:
        lc = lc[~lc['Keterangan'].str.contains('Tgl Penilaian Penilai Independen', na=False)]

    lc['Kategori'] = 'Irrecovable L/C'
    # RENAME NAMA KOLOM
    finallc = lc.rename(columns={
        "Baki Debet": "Baki Debet/Nominal",
        "No L/C": "No Rek/LC/Surat",
        "Jenis L/C": "Jenis Kredit/LC/Garansi/Surat/Fasilitas",
        "Tanggal Keluar": "Tanggal Mulai/Terbit",
        "Tanggal Wan Prestasi": "Tanggal Macet/Wanprestasi",
        "Setoran Jaminan": "Nilai Perolehan/Jaminan/Realisasi",
        "Tujuan L/C": "Tujuan/Jenis Penggunaan",
    })
    
    return finallc

def process_garansi_data(combined_data):
    """Memproses data Garansi dari data gabungan"""
    # Pilih Baris == 'Informasi Debitur' atau 'Jenis Garansi'
    garansi = combined_data[combined_data[0].str.contains('Sistem Layanan Informasi Keuangan|Jenis Garansi', na=False)]
    garansi = pd.DataFrame(garansi)
    lines = garansi[0].apply(lambda x: x.split('\n')).explode().tolist()
    data_list = []
    data_dict = {
        "BANK": None,
        "Baki Debet": None,
        "No Rekening": None,
        "Kualitas": None,
        "Jenis Garansi": None,
        "Tanggal Diterbitkan": None,
        "Tanggal Jatuh Tempo": None,
        "No Akad Awal": None,
        "Tanggal Akad Awal": None,
        "No Akad Akhir": None,
        "Tanggal Akad Akhir": None,
        "Nama Yang Dijamin": None,
        "Keterangan": None,
        "Nama Debitur": None,
        "Nama Group": None
    } 

    nama_debitur = None
    nama_group = None

    for i, line in enumerate(lines):
        if line.strip() == "Nomor Laporan":  # Hanya baris yang persis "Nomor Laporan"
            if i + 2 >= 0:  # Pastikan indeks tidak negatif
                group_line = lines[i + 2].strip()
                nama_group = " ".join(group_line.split()[:3])
                data_dict["Nama Group"] = nama_group
                
        # Baris == "Penyajian informasi debitur pada Sistem Layanan Informasi"
        if "Penyajian informasi debitur pada Sistem Layanan Informasi" in line:
            # Ambil Baris Ke-4 Setelah Teks Tersebut
            if i + 3 < len(lines):
                nama_line = lines[i + 3].strip()  
                # Ambil 5 Kata Pertama
                nama_debitur = " ".join(nama_line.split()[:5])
                data_dict["Nama Debitur"] = nama_debitur

        # NAMA PELAPOR/BANK
        if " - " in line and "Rp" in line:
            bank_part = line.split("Rp")[0].strip()
            data_dict["BANK"] = bank_part

        # BAKI DEBET
            os_part = line.split("Rp")[1].strip()
            os_value = os_part.split(" ")[0]
            data_dict["Baki Debet"] = os_value

        # NO REKENING & KUALITAS
        elif "No Rekening" in line and "Kualitas" in line:
            norek_parts = line.split("No Rekening")[1].strip("Kualitas")[0].strip()
            data_dict["No Rekening"] = norek_parts
            kualitas_parts = line.split("Kualitas")[1].strip()
            data_dict["Kualitas"] = kualitas_parts
        
        # JENIS GARANSI
        elif "Jenis Garansi" in line and "Valuta":
            jenisgar_parts = line.split("Jenis Garansi")[1].strip()
            data_dict["Jenis Garansi"] = jenisgar_parts

        # TANGGAL DITERBITKAN
        elif "Tanggal Diterbitkan" in line and "Plafon":
            tglterbit_parts = line.split("Tanggal Diterbitkan")[1].strip()
            data_dict["Tanggal Diterbitkan"] = tglterbit_parts

        # TANGGAL JATUH TEMPO
        elif "Tanggal Jatuh Tempo" in line and "Tujuan Garansi":
            tgljatem_parts = line.split("Tanggal Jatuh Tempo")[1].strip()
            data_dict["Tanggal Jatuh Tempo"] = tgljatem_parts

        # NO AKAD AWAL
        elif "No Akad Awal" in line and "Setoran Jaminan":
            noakadawal_parts = line.split("No Akad Awal")[1].strip()
            data_dict["No Akad Awal"] = noakadawal_parts

        # TANGGAL AKAD AWAL
        elif "Tanggal Akad Awal" in line and "Tanggal Wan Prestasi":
            tglakadawal_parts = line.split("Tanggal Akad Awal")[1].strip()
            data_dict["Tanggal Akad Awal"] = tglakadawal_parts

        # NO AKAD AKHIR
        elif "No Akad Akhir" in line and "Kondisi":
            noakadakhir_parts = line.split("No Akad Akhir")[1].strip()
            data_dict["No Akad Akhir"] = noakadakhir_parts

        # TANGGAL AKAD AKHIR
        elif "Tanggal Akad Akhir" in line and "Tanggal Kondisi":
            tglakadakhir_parts = line.split("Tanggal Akad Akhir")[1].strip()
            data_dict["Tanggal Akad Akhir"] = tglakadakhir_parts
        
        # NAMA YANG DIJAMIN
        elif "Nama Yang Dijamin" in line:
            namajamin_parts = line.split("Nama Yang Dijamin")[1].strip()
            data_dict["Nama Yang Dijamin"] = namajamin_parts

        # KETERANGAN
        elif "Keterangan" in line:
            # Cek apakah baris sebelumnya mengandung Nama Yang Dijamin
            if i > 0 and ("Nama Yang Dijamin" in lines[i-1]):
                keterangan_parts = line.split("Keterangan")[1].strip()
                data_dict["Keterangan"] = keterangan_parts
                
                # MENYIMPAN DICTIONARY KEDALAM LIST
                data_list.append(data_dict.copy())

    garansi = pd.DataFrame(data_list)

    if 'Jenis Garansi' in garansi.columns:
        garansi = garansi.dropna(subset=['Jenis Garansi'])
    else:
        garansi['Jenis Garansi'] = pd.NA

    if not garansi.empty and 'Jenis Garansi' in garansi.columns:
        garansi[['Jenis Garansi', 'Valuta']] = garansi['Jenis Garansi'].apply(
            lambda x: pd.Series(x.split('Valuta', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not garansi.empty and 'Tanggal Diterbitkan' in garansi.columns:
        garansi[['Tanggal Diterbitkan', 'Plafon']] = garansi['Tanggal Diterbitkan'].apply(
            lambda x: pd.Series(x.split('Plafon', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not garansi.empty and 'Tanggal Jatuh Tempo' in garansi.columns:
        garansi[['Tanggal Jatuh Tempo', 'Tujuan Garansi']] = garansi['Tanggal Jatuh Tempo'].apply(
            lambda x: pd.Series(x.split('Tujuan Garansi', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not garansi.empty and 'No Akad Awal' in garansi.columns:
        garansi[['No Akad Awal', 'Setoran Jaminan']] = garansi['No Akad Awal'].apply(
            lambda x: pd.Series(x.split('Setoran Jaminan', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not garansi.empty and 'Tanggal Akad Awal' in garansi.columns:
        garansi[['Tanggal Akad Awal', 'Tanggal Wan Prestasi']] = garansi['Tanggal Akad Awal'].apply(
            lambda x: pd.Series(x.split('Tanggal Wan Prestasi', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not garansi.empty and 'No Akad Akhir' in garansi.columns:
        garansi[['No Akad Akhir', 'Kondisi']] = garansi['No Akad Akhir'].apply(
            lambda x: pd.Series(x.split('Kondisi', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not garansi.empty and 'Tanggal Akad Akhir' in garansi.columns:
        garansi[['Tanggal Akad Akhir', 'Tanggal Kondisi']] = garansi['Tanggal Akad Akhir'].apply(
            lambda x: pd.Series(x.split('Tanggal Kondisi', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not garansi.empty and 'Keterangan' in garansi.columns:
        garansi = garansi[~garansi['Keterangan'].str.contains('Tgl Penilaian Penilai Independen', na=False)]

    garansi['Kategori'] = 'Garansi'
    finalgaransi = garansi.rename(columns={
        "Baki Debet": "Baki Debet/Nominal",
        "No Rekening": "No Rek/LC/Surat",
        "Jenis Garansi": "Jenis Kredit/LC/Garansi/Surat/Fasilitas",
        "Tanggal Diterbitkan": "Tanggal Mulai/Terbit",
        "Tanggal Wan Prestasi": "Tanggal Macet/Wanprestasi",
        "Setoran Jaminan": "Nilai Perolehan/Jaminan/Realisasi",
        "Tujuan Garansi": "Tujuan/Jenis Penggunaan",
    })
    
    return finalgaransi

def process_surat_data(combined_data):
    """Memproses data Surat Berharga dari data gabungan"""
    # Pilih Baris == 'Informasi Debitur' atau 'Jenis Surat Berharga'
    surat = combined_data[combined_data[0].str.contains('Sistem Layanan Informasi Keuangan|Jenis Surat Berharga', na=False)]
    surat = pd.DataFrame(surat)
    lines = surat[0].apply(lambda x: x.split('\n')).explode().tolist()
    data_list = []
    data_dict = {
        "BANK": None,
        "Baki Debet": None,
        "No Surat Berharga": None,
        "Kualitas": None,
        "Jenis Surat Berharga": None,
        "Sovereign Rate": None,
        "Listing": None,
        "Peringkat Surat Berharga": None,
        "Tujuan Kepemilikan": None,
        "Tanggal Terbit": None,
        "Tanggal Jatuh Tempo": None,
        "Suku Bunga/Imbalan": None,
        "Kode Valuta": None,
        "Keterangan": None,
        "Nama Debitur": None,
        "Nama Group": None
    } 
    
    nama_debitur = None
    nama_group = None

    for i, line in enumerate(lines):
        if line.strip() == "Nomor Laporan":  # Hanya baris yang persis "Nomor Laporan"
            if i + 2 >= 0:  # Pastikan indeks tidak negatif
                group_line = lines[i + 2].strip()
                nama_group = " ".join(group_line.split()[:3])
                data_dict["Nama Group"] = nama_group
                
        # Baris == "Penyajian informasi debitur pada Sistem Layanan Informasi"
        if "Penyajian informasi debitur pada Sistem Layanan Informasi" in line:
            # Ambil Baris Ke-4 Setelah Teks Tersebut
            if i + 3 < len(lines):
                nama_line = lines[i + 3].strip()  
                # Ambil 5 Kata Pertama
                nama_debitur = " ".join(nama_line.split()[:5])
                data_dict["Nama Debitur"] = nama_debitur

        # NAMA PELAPOR/BANK
        if " - " in line and "Rp" in line:
            bank_part = line.split("Rp")[0].strip()
            data_dict["BANK"] = bank_part

        # BAKI DEBET
            os_part = line.split("Rp")[1].strip()
            os_value = os_part.split(" ")[0]
            data_dict["Baki Debet"] = os_value

        # NO SURAT BERHARGA & KUALITAS
        elif "No Surat Berharga" in line and "Kualitas" in line:
            nosurat_parts = line.split("No Surat Berharga")[1].strip("Kualitas")[0].strip()
            data_dict["No Surat Berharga"] = nosurat_parts
            kualitas_parts = line.split("Kualitas")[1].strip()
            data_dict["Kualitas"] = kualitas_parts
        
        # JENIS SURAT BERHARGA
        elif "Jenis Surat Berharga" in line and "Jumlah Hari Tunggakan":
            jenissurat_parts = line.split("Jenis Surat Berharga")[1].strip()
            data_dict["Jenis Surat Berharga"] = jenissurat_parts

        # SOVEREIGN RATE
        elif "Sovereign Rate" in line and "Nilai Dalam Mata Uang Asal":
            sovrate_parts = line.split("Sovereign Rate")[1].strip()
            data_dict["Sovereign Rate"] = sovrate_parts

        # LISTING
        elif "Listing" in line and "Nilai Pasar":
            listing_parts = line.split("Listing")[1].strip()
            data_dict["Listing"] = listing_parts

        # PERINGKAT SURAT BERHARGA
        elif "Peringkat Surat Berharga" in line and "Nilai Perolehan":
            persuber_parts = line.split("Peringkat Surat Berharga")[1].strip()
            data_dict["Peringkat Surat Berharga"] = persuber_parts

        # TUJUAN KEPEMILIKAN
        elif "Tujuan Kepemilikan" in line and "Tunggakan":
            tujuanpe_parts = line.split("Tujuan Kepemilikan")[1].strip()
            data_dict["Tujuan Kepemilikan"] = tujuanpe_parts

        # TANGGAL TERBIT
        elif "Tanggal Terbit" in line and "Tanggal Macet":
            tglterbit_parts = line.split("Tanggal Terbit")[1].strip()
            data_dict["Tanggal Terbit"] = tglterbit_parts

        # TANGGAL JATUH TEMPO
        elif "Tanggal Jatuh Tempo" in line and "Sebab Macet":
            tgljatem_parts = line.split("Tanggal Jatuh Tempo")[1].strip()
            data_dict["Tanggal Jatuh Tempo"] = tgljatem_parts
        
        # SUKA BUNGA/IMBALAN
        elif "Suku Bunga/Imbalan" in line and "Kondisi":
            sukubunga_parts = line.split("Suku Bunga/Imbalan")[1].strip()
            data_dict["Suku Bunga/Imbalan"] = sukubunga_parts

        # KODE VALUTA
        elif "Kode Valuta" in line and "Tanggal Kondisi":
            kodevaluta_parts = line.split("Kode Valuta")[1].strip()
            data_dict["Kode Valuta"] = kodevaluta_parts

        # KETERANGAN
        elif "Keterangan" in line:
            # Cek apakah baris sebelumnya mengandung Kode Valuta atau Tanggal Kondisi
            if i > 0 and ("Kode Valuta" in lines[i-1] or "Tanggal Kondisi" in lines[i-1]):
                keterangan_parts = line.split("Keterangan")[1].strip()
                data_dict["Keterangan"] = keterangan_parts
                
                # MENYIMPAN DICTIONARY KEDALAM LIST
                data_list.append(data_dict.copy())

    surat = pd.DataFrame(data_list)

    if 'Jenis Surat Berharga' in surat.columns:
        surat = surat.dropna(subset=['Jenis Surat Berharga'])
    else:
        surat['Jenis Surat Berharga'] = pd.NA

    if not surat.empty and 'Jenis Surat Berharga' in surat.columns:
        surat[['Jenis Surat Berharga', 'Jumlah Hari Tunggakan']] = surat['Jenis Surat Berharga'].apply(
            lambda x: pd.Series(x.split('Jumlah Hari Tunggakan', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not surat.empty and 'Sovereign Rate' in surat.columns:
        surat[['Sovereign Rate', 'Nilai Dalam Mata Uang Asal']] = surat['Sovereign Rate'].apply(
            lambda x: pd.Series(x.split('Nilai Dalam Mata Uang Asal', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not surat.empty and 'Listing' in surat.columns:
        surat[['Listing', 'Nilai Pasar']] = surat['Listing'].apply(
            lambda x: pd.Series(x.split('Nilai Pasar', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not surat.empty and 'Peringkat Surat Berharga' in surat.columns:
        surat[['Peringkat Surat Berharga', 'Nilai Perolehan']] = surat['Peringkat Surat Berharga'].apply(
            lambda x: pd.Series(x.split('Nilai Perolehan', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not surat.empty and 'Tujuan Kepemilikan' in surat.columns:
        surat[['Tujuan Kepemilikan', 'Tunggakan']] = surat['Tujuan Kepemilikan'].apply(
            lambda x: pd.Series(x.split('Tunggakan', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not surat.empty and 'Tanggal Terbit' in surat.columns:
        surat[['Tanggal Terbit', 'Tanggal Macet']] = surat['Tanggal Terbit'].apply(
            lambda x: pd.Series(x.split('Tanggal Macet', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not surat.empty and 'Tanggal Jatuh Tempo' in surat.columns:
        surat[['Tanggal Jatuh Tempo', 'Sebab Macet']] = surat['Tanggal Jatuh Tempo'].apply(
            lambda x: pd.Series(x.split('Sebab Macet', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not surat.empty and 'Suku Bunga/Imbalan' in surat.columns:
        surat[['Suku Bunga/Imbalan', 'Kondisi']] = surat['Suku Bunga/Imbalan'].apply(
            lambda x: pd.Series(x.split('Kondisi', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not surat.empty and 'Kode Valuta' in surat.columns:
        surat[['Kode Valuta', 'Tanggal Kondisi']] = surat['Kode Valuta'].apply(
            lambda x: pd.Series(x.split('Tanggal Kondisi', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    # Fungsi untuk membersihkan persentase
    def clean_percentage(value):
        if pd.isnull(value):
            return value 
        value = str(value).replace('%', '').strip()
        try:
            return float(value)
        except ValueError:
            return None

    if not surat.empty and 'Suku Bunga/Imbalan' in surat.columns:
        surat['Suku Bunga/Imbalan'] = surat['Suku Bunga/Imbalan'].apply(clean_percentage)

    if not surat.empty and 'Keterangan' in surat.columns:
        surat = surat[~surat['Keterangan'].str.contains('Tgl Penilaian Penilai Independen', na=False)]

    surat['Kategori'] = 'Surat Berharga'
    # RENAME NAMA KOLOM
    finalsurat = surat.rename(columns={
        "Baki Debet": "Baki Debet/Nominal",
        "No Surat Berharga": "No Rek/LC/Surat",
        "Jenis Surat Berharga": "Jenis Kredit/LC/Garansi/Surat/Fasilitas",
        "Tanggal Terbit": "Tanggal Mulai/Terbit",
        "Tanggal Macet": "Tanggal Macet/Wanprestasi",
        "Kode Valuta": "Valuta",
        "Nilai Perolehan": "Nilai Perolehan/Jaminan/Realisasi",
        "Tunggakan": "Tunggakan Pokok",
        "Tujuan Kepemilikan": "Tujuan/Jenis Penggunaan",
        "Nilai Pasar": "Nilai Pasar/Proyek",
    })
    
    return finalsurat

def process_fasilitas_data(combined_data):
    """Memproses data Fasilitas Lain dari data gabungan"""
    # Pilih Baris == 'Informasi Debitur' atau 'Jenis Fasilitas'
    fasilitas = combined_data[combined_data[0].str.contains('Sistem Layanan Informasi Keuangan|Jenis Fasilitas', na=False)]
    fasilitas = pd.DataFrame(fasilitas)
    lines = fasilitas[0].apply(lambda x: x.split('\n')).explode().tolist()
    data_list = []
    data_dict = {
        "BANK": None,
        "Baki Debet": None,
        "No Rekening": None,
        "Kualitas": None,
        "Jenis Fasilitas": None,
        "Tanggal Mulai": None,
        "Tanggal Jatuh Tempo": None,
        "Valuta": None,
        "Nilai Dalam Mata Uang Asal": None,
        "Suku Bunga/Imbalan": None,
        "Keterangan": None,
        "Nama Debitur": None,
        "Nama Group": None
    } 

    nama_debitur = None
    nama_group = None

    for i, line in enumerate(lines):
        if line.strip() == "Nomor Laporan":  # Hanya baris yang persis "Nomor Laporan"
            if i + 2 >= 0:  # Pastikan indeks tidak negatif
                group_line = lines[i + 2].strip()
                nama_group = " ".join(group_line.split()[:3])
                data_dict["Nama Group"] = nama_group
                
        # Baris == "Penyajian informasi debitur pada Sistem Layanan Informasi"
        if "Penyajian informasi debitur pada Sistem Layanan Informasi" in line:
            # Ambil Baris Ke-4 Setelah Teks Tersebut
            if i + 3 < len(lines):
                nama_line = lines[i + 3].strip()  
                # Ambil 5 Kata Pertama
                nama_debitur = " ".join(nama_line.split()[:5])
                data_dict["Nama Debitur"] = nama_debitur

        # NAMA PELAPOR/BANK
        if " - " in line and "Rp" in line:
            bank_part = line.split("Rp")[0].strip()
            data_dict["BANK"] = bank_part

        # BAKI DEBET
            os_part = line.split("Rp")[1].strip()
            os_value = os_part.split(" ")[0]
            data_dict["Baki Debet"] = os_value

        # NO REKENING & KUALITAS
        elif "No Rekening" in line and "Kualitas" in line:
            norek_parts = line.split("No Rekening")[1].strip("Kualitas")[0].strip()
            data_dict["No Rekening"] = norek_parts
            kualitas_parts = line.split("Kualitas")[1].strip()
            data_dict["Kualitas"] = kualitas_parts
        
        # JENIS FASILITAS
        elif "Jenis Fasilitas" in line and "Jumlah Hari Tunggakan":
            jenisfas_parts = line.split("Jenis Fasilitas")[1].strip()
            data_dict["Jenis Fasilitas"] = jenisfas_parts

        # TANGGAL MULAI
        elif "Tanggal Mulai" in line and "Tanggal Macet":
            tglmulai_parts = line.split("Tanggal Mulai")[1].strip()
            data_dict["Tanggal Mulai"] = tglmulai_parts

        # TANGGAL JATUH TEMPO
        elif "Tanggal Jatuh Tempo" in line and "Sebab Macet":
            tgljatem_parts = line.split("Tanggal Jatuh Tempo")[1].strip()
            data_dict["Tanggal Jatuh Tempo"] = tgljatem_parts

        # VALUTA
        elif "Valuta" in line and "Tunggakan":
            valuta_parts = line.split("Valuta")[1].strip()
            data_dict["Valuta"] = valuta_parts

        # NILAI DALAM MATA UANG ASAL
        elif "Nilai Dalam Mata Uang Asal" in line and "Kondisi":
            uangasal_parts = line.split("Nilai Dalam Mata Uang Asal")[1].strip()
            data_dict["Nilai Dalam Mata Uang Asal"] = uangasal_parts
        
        # SUKA BUNGA/IMBALAN
        elif "Suku Bunga/Imbalan" in line and "Tanggal Kondisi":
            sukubunga_parts = line.split("Suku Bunga/Imbalan")[1].strip()
            data_dict["Suku Bunga/Imbalan"] = sukubunga_parts

        # KETERANGAN
        elif "Keterangan" in line:
            # Cek apakah baris sebelumnya mengandung Suku Bunga/Imbalan atau Tanggal Kondisi
            if i > 0 and ("Suku Bunga/Imbalan" in lines[i-1] or "Tanggal Kondisi" in lines[i-1]):
                keterangan_parts = line.split("Keterangan")[1].strip()
                data_dict["Keterangan"] = keterangan_parts
                
                # MENYIMPAN DICTIONARY KEDALAM LIST
                data_list.append(data_dict.copy())

    fasilitas = pd.DataFrame(data_list)

    if 'Jenis Fasilitas' in fasilitas.columns:
        fasilitas = fasilitas.dropna(subset=['Jenis Fasilitas'])
    else:
        fasilitas['Jenis Fasilitas'] = pd.NA

    if not fasilitas.empty and 'Jenis Fasilitas' in fasilitas.columns:
        fasilitas[['Jenis Fasilitas', 'Jumlah Hari Tunggakan']] = fasilitas['Jenis Fasilitas'].apply(
            lambda x: pd.Series(x.split('Jumlah Hari Tunggakan', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not fasilitas.empty and 'Tanggal Mulai' in fasilitas.columns:
        fasilitas[['Tanggal Mulai', 'Tanggal Macet']] = fasilitas['Tanggal Mulai'].apply(
            lambda x: pd.Series(x.split('Tanggal Macet', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not fasilitas.empty and 'Tanggal Jatuh Tempo' in fasilitas.columns:
        fasilitas[['Tanggal Jatuh Tempo', 'Sebab Macet']] = fasilitas['Tanggal Jatuh Tempo'].apply(
            lambda x: pd.Series(x.split('Sebab Macet', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not fasilitas.empty and 'Valuta' in fasilitas.columns:
        fasilitas[['Valuta', 'Tunggakan']] = fasilitas['Valuta'].apply(
            lambda x: pd.Series(x.split('Tunggakan', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not fasilitas.empty and 'Nilai Dalam Mata Uang Asal' in fasilitas.columns:
        fasilitas[['Nilai Dalam Mata Uang Asal', 'Kondisi']] = fasilitas['Nilai Dalam Mata Uang Asal'].apply(
            lambda x: pd.Series(x.split('Kondisi', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    if not fasilitas.empty and 'Suku Bunga/Imbalan' in fasilitas.columns:
        fasilitas[['Suku Bunga/Imbalan', 'Tanggal Kondisi']] = fasilitas['Suku Bunga/Imbalan'].apply(
            lambda x: pd.Series(x.split('Tanggal Kondisi', 1)) if pd.notna(x) else pd.Series([None, None])
        )

    # Fungsi untuk membersihkan persentase
    def clean_percentage(value):
        if pd.isnull(value):
            return value 
        value = str(value).replace('%', '').strip()
        try:
            return float(value)
        except ValueError:
            return None

    if not fasilitas.empty and 'Suku Bunga/Imbalan' in fasilitas.columns:
        fasilitas['Suku Bunga/Imbalan'] = fasilitas['Suku Bunga/Imbalan'].apply(clean_percentage)

    if not fasilitas.empty and 'Keterangan' in fasilitas.columns:
        fasilitas = fasilitas[~fasilitas['Keterangan'].str.contains('Tgl Penilaian Penilai Independen', na=False)]

    fasilitas['Kategori'] = 'Fasilitas Lain'
    finalfasilitas = fasilitas.rename(columns={
        "Baki Debet": "Baki Debet/Nominal",
        "No Rekening": "No Rek/LC/Surat",
        "Jenis Fasilitas": "Jenis Kredit/LC/Garansi/Surat/Fasilitas",
        "Tanggal Mulai": "Tanggal Mulai/Terbit",
        "Tanggal Macet": "Tanggal Macet/Wanprestasi",
        "Nilai Perolehan": "Nilai Perolehan/Jaminan/Realisasi",
        "Tunggakan": "Tunggakan Pokok",
    })
    
    return finalfasilitas

# Fungsi untuk membersihkan persentase
def clean_percentage(value):
    if pd.isnull(value):
        return value 
    value = str(value).replace('%', '').strip()
    try:
        return float(value) / 100
    except ValueError:
        return None

# Fungsi untuk membersihkan data numerik
def clean_numeric_data(value):
    value = str(value).replace('Rp', '')  # Menghapus teks "RP"
    value = value.split(',')[0]  # Menghapus karakter koma dan nilai setelah koma
    value = value.replace('.', '')  # Menghapus karakter titik (.)
    return value

# Mapping untuk typo bulan
typo_mapping = {
    'amgoursttiusas': 'Agustus',
    'nmoovretimsabsier': 'November',
    'jmuloi': 'Juli',
    'smepotretimsabsier': 'September',
    'omkotortbisears': 'Oktober',
    'mmeoi': 'Mei',
    'jmanourtaisrai': 'Januari',
    'meborrutiasaris': 'Februari',
    'amporilr': 'April',
    'mmaorerttis': 'Maret',
    'dmeosertmisabseir': 'Desember',
    'jmunoi': 'Juni'
}

month_translation = {
    'Januari': 'January',
    'Februari': 'February',
    'Maret': 'March',
    'April': 'April',
    'Mei': 'May',
    'Juni': 'June',
    'Juli': 'July',
    'Agustus': 'August',
    'September': 'September',
    'Oktober': 'October',
    'November': 'November',
    'Desember': 'December'
}

def process_date_string(date_value):
    # Handle missing values
    if pd.isna(date_value) or str(date_value).strip() in ['', 'None', 'NaT']:
        return None
    
    date_str = str(date_value).strip()
    
    try:
        # 1. Coba parsing langsung dulu
        parsed = pd.to_datetime(date_str, dayfirst=True, errors='coerce')
        if not pd.isna(parsed):
            return parsed
        
        # 2. Perbaiki typo bulan
        for typo, correct in typo_mapping.items():
            if typo in date_str.lower():
                date_str = date_str.lower().replace(typo, correct.lower())
                break
        
        # 3. Pisahkan komponen tanggal
        parts = re.split(r'[\s/-]+', date_str)
        if len(parts) != 3:
            return None
        
        day, month, year = parts
        
        # 4. Bersihkan hari dan tahun
        day = re.sub(r'[^\d]', '', day).zfill(2)
        year = re.sub(r'[^\d]', '', year)
        
        # Handle tahun 2 digit
        if len(year) == 2:
            year = f'20{year}' if int(year) < 50 else f'19{year}'
        
        # 5. Terjemahkan bulan ke Inggris
        month_corrected = month_translation.get(month.capitalize(), month)
        
        # 6. Gabungkan dan parse
        formatted_date = f"{day} {month_corrected} {year}"
        parsed = pd.to_datetime(formatted_date, format='%d %B %Y', errors='coerce')
        
        return parsed if not pd.isna(parsed) else None
    
    except Exception as e:
        return None

def combine_and_clean_data(finalkredit, finallc, finalgaransi, finalsurat, finalfasilitas):
    """Menggabungkan dan membersihkan semua data"""
    
    # GABUNG SEMUA DATA
    st.info("🔄 Menggabungkan semua data...")
    Gabungan1 = [finalkredit, finallc, finalgaransi, finalsurat, finalfasilitas]
    Gabungan2 = pd.concat(Gabungan1, axis=0, ignore_index=True, sort=False)
    
    st.success(f"✅ Data berhasil digabungkan. Total {len(Gabungan2)} baris")

    # Daftar kolom yang ingin diambil
    kolom_dipilih = ['Nama Group','Nama Debitur','BANK','CABANG','Kategori',
                     'Tujuan/Jenis Penggunaan','Plafon',
                     'Baki Debet/Nominal','Suku Bunga/Imbalan','Kualitas',
                     'Valuta','Kondisi','Tanggal Kondisi','Tanggal Jatuh Tempo',
                     'Jenis Kredit/LC/Garansi/Surat/Fasilitas','Tanggal Mulai/Terbit',
                     'Nilai Perolehan/Jaminan/Realisasi','Tunggakan Pokok',
                     'Tunggakan Bunga','Frekuensi Tunggakan','Jumlah Hari Tunggakan',
                     'Tanggal Macet/Wanprestasi','Sebab Macet','Keterangan',
                     'No Akad Awal','Tanggal Akad Awal','No Akad Akhir',
                     'Tanggal Akad Akhir','Frekuensi Restrukturisasi',
                     'Cara Restrukturisasi','Tanggal Restrukturisasi Akhir',
                     'Tanggal Awal Kredit','Plafon Awal','Sifat Kredit/Pembiayaan',
                     'Akad Kredit/Pembiayaan','Frekuensi Perpanjangan Kredit/',
                     'Kategori Debitur','Sektor Ekonomi','Kab/Kota Lokasi Proyek',
                     'Kredit Program Pemerintah','Jenis Suku Bunga/Imbalan',
                     'Denda','Bank Beneficiary','Nama Yang Dijamin','Sovereign Rate',
                     'Listing','Peringkat Surat Berharga','Nilai Dalam Mata Uang Asal',
                     'No Rek/LC/Surat','Nilai Pasar/Proyek']

    # Ambil hanya kolom yang ada di DataFrame
    kolom_tersedia = [kol for kol in kolom_dipilih if kol in Gabungan2.columns]
    
    # Buat DataFrame baru hanya dengan kolom yang tersedia
    Gabungan3 = Gabungan2[kolom_tersedia]
    
    st.info(f"📊 Kolom yang tersedia: {len(kolom_tersedia)} dari {len(kolom_dipilih)}")

    # PROSES BANK DAN CABANG
    st.info("🏦 Memproses data bank dan cabang...")
    try:
        # Load kode bank (asumsi file ada di direktori yang sama)
        kodebank_path = "ZZ Data Kode Bank.xlsx"
        if os.path.exists(kodebank_path):
            kodebank = pd.read_excel(kodebank_path)
            
            if not Gabungan3.empty and 'BANK' in Gabungan3.columns and not kodebank.empty and 'KETERANGAN' in kodebank.columns:
                def pisahkan_bank_cabang(row, kodebank):
                    for keterangan in kodebank['KETERANGAN']:
                        if keterangan in row:
                            cabang = row.replace(keterangan, '').strip()
                            return pd.Series([keterangan, cabang])
                    return pd.Series([row, '']) 

                # APPLY FUNCTION
                Gabungan3[['BANK', 'CABANG']] = Gabungan3['BANK'].apply(lambda x: pisahkan_bank_cabang(x, kodebank))

            if not Gabungan3.empty and 'BANK' in Gabungan3.columns:
                Gabungan3[['KODE BANK', 'NAMA BANK']] = Gabungan3['BANK'].str.split(' - ', n=1, expand=True)

            if not Gabungan3.empty and 'BANK' in Gabungan3.columns and not kodebank.empty and 'KETERANGAN' in kodebank.columns and 'NAMA BANK' in kodebank.columns:
                # Membuat dict untuk mapping dari kodebank
                mapping_dict = dict(zip(kodebank['KETERANGAN'], kodebank['NAMA BANK']))
                # Replace nilai kolom 'BANK'
                Gabungan3['BANK'] = Gabungan3['BANK'].map(mapping_dict)
                
            st.success("✅ Data bank berhasil diproses")
        else:
            st.warning("⚠️ File kode bank tidak ditemukan, melanjutkan tanpa pemrosesan bank")
    except Exception as e:
        st.warning(f"⚠️ Gagal memproses data bank: {str(e)}")

    # CLEANING NAMA DEBITUR DAN GROUP
    st.info("👤 Membersihkan nama debitur dan group...")
    if not Gabungan3.empty and 'Nama Debitur' in Gabungan3.columns and 'Nama Group' in Gabungan3.columns:
        Gabungan3["Nama Debitur"] = (
            Gabungan3["Nama Debitur"]
            .str.replace(r'NIK\s?/.*|NPWP\s?/.*', '', regex=True)
            .str.replace(r'\s*\d+.*$', '', regex=True)
            .str.replace(r'(LAKI|PEREMPUAN).*$', '', regex=True)
            .str.strip()
        )

        Gabungan3["Nama Group"] = (
            Gabungan3["Nama Group"]
            .str.replace(r'(?i)(posisi|laki|perempuan).*$', '', regex=True)
            .str.strip()
        )
        st.success("✅ Nama debitur dan group berhasil dibersihkan")

    # CLEANING DATA NUMERIK
    st.info("🔢 Membersihkan data numerik...")
    if not Gabungan3.empty:
        columns_to_clean = ['Plafon','Baki Debet/Nominal','Nilai Perolehan/Jaminan/Realisasi','Tunggakan Pokok',
                            'Tunggakan Bunga','Plafon Awal','Denda','Nilai Pasar/Proyek']
        columns_to_clean = [col for col in columns_to_clean if col in Gabungan3.columns]

        # Apply ke kolom
        for col in columns_to_clean:
            Gabungan3[col] = Gabungan3[col].apply(clean_numeric_data)
        st.success(f"✅ {len(columns_to_clean)} kolom numerik berhasil dibersihkan")

    # CLEANING TANGGAL
    st.info("📅 Memproses data tanggal...")
    date_columns = ['Tanggal Kondisi','Tanggal Jatuh Tempo','Tanggal Mulai/Terbit','Tanggal Macet/Wanprestasi',
                    'Tanggal Akad Awal','Tanggal Akad Akhir','Tanggal Restrukturisasi Akhir','Tanggal Awal Kredit']

    date_success_count = 0
    for col in date_columns:
        if col in Gabungan3.columns:
            Gabungan3[col] = Gabungan3[col].apply(process_date_string)
            success_rate = (1 - Gabungan3[col].isna().mean()) * 100
            if success_rate > 0:
                date_success_count += 1

    st.success(f"✅ {date_success_count} kolom tanggal berhasil diproses")

    # CLEANING PERSENTASE
    st.info("📊 Membersihkan data persentase...")
    if not Gabungan3.empty and 'Suku Bunga/Imbalan' in Gabungan3.columns:
        Gabungan3['Suku Bunga/Imbalan'] = Gabungan3['Suku Bunga/Imbalan'].apply(clean_percentage)
        st.success("✅ Data suku bunga/imbalan berhasil dibersihkan")

    # KONVERSI TIPE DATA
    st.info("🔄 Mengkonversi tipe data...")
    for col in Gabungan3.select_dtypes(include=['object']):
        try:
            Gabungan3[col] = pd.to_numeric(Gabungan3[col], errors='ignore')
        except ValueError:
            pass
    st.success("✅ Konversi tipe data selesai")

    # FINAL SELECTION
    kolom_tersedia_final = [kol for kol in kolom_dipilih if kol in Gabungan3.columns]
    finaldata = Gabungan3[kolom_tersedia_final]
    
    st.success(f"🎉 Cleaning data selesai! Final data shape: {finaldata.shape}")
    
    return finaldata

def save_to_excel_formatted(df, filename):
    """Menyimpan DataFrame ke Excel dengan formatting"""
    # Buat workbook baru
    wb = Workbook()
    ws = wb.active
    
    # Style border putih yang lebih tebal (medium)
    white_border = Border(
        left=Side(style='medium', color='FFFFFF'),
        right=Side(style='medium', color='FFFFFF'),
        top=Side(style='medium', color='FFFFFF'),
        bottom=Side(style='medium', color='FFFFFF')
    )
    
    # Konversi DataFrame ke rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = white_border
    
    # Format header - navy dengan font putih
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = white_border
    
    # Format baris zigzag putih abu-abu
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.row % 2 == 0:  # Baris genap
                cell.fill = white_fill
            else:  # Baris ganjil
                cell.fill = grey_fill
            cell.border = white_border
    
    # Sesuaikan lebar kolom otomatis
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Simpan file
    wb.save(filename)

def main():
    if uploaded_files:
        st.markdown('<div class="sub-header">📊 File yang Diupload</div>', unsafe_allow_html=True)
        st.write(f"Jumlah file PDF: {len(uploaded_files)}")
        
        for i, file in enumerate(uploaded_files):
            st.write(f"{i+1}. {file.name}")
        
        # Tombol untuk memulai proses
        if st.button("🚀 Mulai Konversi ke Excel", type="primary"):
            with st.spinner("Sedang memproses file..."):
                # Step 1: Konversi PDF ke JSON
                st.markdown('<div class="sub-header">📝 Langkah 1: Konversi PDF ke JSON</div>', unsafe_allow_html=True)
                
                json_files = []
                temp_dir = tempfile.mkdtemp()
                
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                for i, uploaded_file in enumerate(uploaded_files):
                    status_text.text(f"Memproses {uploaded_file.name}...")
                    
                    # Simpan file PDF sementara
                    pdf_path = os.path.join(temp_dir, uploaded_file.name)
                    with open(pdf_path, "wb") as f:
                        f.write(uploaded_file.getbuffer())
                    
                    # Konversi ke JSON
                    json_filename = os.path.splitext(uploaded_file.name)[0] + ".json"
                    json_path = os.path.join(temp_dir, json_filename)
                    
                    pdf_to_json(pdf_path, json_path)
                    json_files.append(json_path)
                    
                    progress_bar.progress((i + 1) / len(uploaded_files))
                
                status_text.text("✅ Semua file PDF berhasil dikonversi ke JSON")
                st.markdown('<div class="success-box">✅ Konversi PDF ke JSON selesai!</div>', unsafe_allow_html=True)
                
                # Step 2: Baca file JSON
                st.markdown('<div class="sub-header">📖 Langkah 2: Membaca File JSON</div>', unsafe_allow_html=True)
                
                combined_data = read_json_files(json_files)
                
                if combined_data is not None:
                    st.markdown('<div class="success-box">✅ Pembacaan file JSON berhasil!</div>', unsafe_allow_html=True)
                    st.write(f"Total data yang digabungkan: {len(combined_data)} baris")
                    
                    # Step 3: Proses semua jenis data
                    st.markdown('<div class="sub-header">🔧 Langkah 3: Memproses Semua Jenis Data</div>', unsafe_allow_html=True)
                    
                    # Proses masing-masing jenis data
                    with st.spinner("Memproses data Kredit..."):
                        kredit_data = process_kredit_data(combined_data)
                    
                    with st.spinner("Memproses data LC..."):
                        lc_data = process_lc_data(combined_data)
                    
                    with st.spinner("Memproses data Garansi..."):
                        garansi_data = process_garansi_data(combined_data)
                    
                    with st.spinner("Memproses data Surat..."):
                        surat_data = process_surat_data(combined_data)
                    
                    with st.spinner("Memproses data Fasilitas..."):
                        fasilitas_data = process_fasilitas_data(combined_data)
                    
                    # Tampilkan summary data
                    st.markdown("### 📈 Summary Data yang Diproses")
                    col1, col2, col3, col4, col5 = st.columns(5)
                    
                    with col1:
                        st.metric("Kredit", f"{len(kredit_data)}" if not kredit_data.empty else "0")
                    with col2:
                        st.metric("LC", f"{len(lc_data)}" if not lc_data.empty else "0")
                    with col3:
                        st.metric("Garansi", f"{len(garansi_data)}" if not garansi_data.empty else "0")
                    with col4:
                        st.metric("Surat", f"{len(surat_data)}" if not surat_data.empty else "0")
                    with col5:
                        st.metric("Fasilitas", f"{len(fasilitas_data)}" if not fasilitas_data.empty else "0")
                    
                    # Step 4: Cleaning final data
                    st.markdown('<div class="sub-header">✨ Langkah 4: Cleaning Data Final</div>', unsafe_allow_html=True)
                    
                    with st.spinner("Menggabungkan dan membersihkan data..."):
                        final_data = combine_and_clean_data(kredit_data, lc_data, garansi_data, surat_data, fasilitas_data)
                    
                    # Tampilkan preview data
                    st.markdown('<div class="sub-header">👁️ Preview Data</div>', unsafe_allow_html=True)
                    
                    # Tampilkan statistik data
                    st.write(f"**Shape data final:** {final_data.shape}")
                    st.write(f"**Total records:** {len(final_data)}")
                    st.write(f"**Total kolom:** {len(final_data.columns)}")
                    
                    # Tampilkan preview tabel
                    st.dataframe(final_data.head(10))
                    
                    # Step 5: Download Excel
                    st.markdown('<div class="sub-header">💾 Langkah 5: Download File Excel</div>', unsafe_allow_html=True)
                    
                    # Simpan ke Excel
                    with st.spinner("Menyimpan ke format Excel..."):
                        excel_buffer = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                        save_to_excel_formatted(final_data, excel_buffer.name)
                        excel_buffer.close()

                        # Baca file Excel untuk download
                        with open(excel_buffer.name, "rb") as f:
                            excel_data = f.read()
                    
                    # Tombol download
                    st.download_button(
                        label="📥 Download Excel File",
                        data=excel_data,
                        file_name="SLIK_Report_Converted.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )
                    
                    st.markdown('<div class="success-box">✅ Proses selesai! File Excel siap diunduh.</div>', unsafe_allow_html=True)
                    
                    # Tampilkan informasi file
                    file_size = len(excel_data) / 1024 / 1024  # Convert to MB
                    st.info(f"**Ukuran file:** {file_size:.2f} MB")
                    
                    # Bersihkan file temporary
                    try:
                        os.unlink(excel_buffer.name)
                    except PermissionError as e:
                        st.warning(f"Tidak bisa menghapus file sementara: {e}")
                    
                else:
                    st.error("❌ Gagal membaca file JSON.")
                
                # Bersihkan directory temporary
                for json_file in json_files:
                    if os.path.exists(json_file):
                        os.unlink(json_file)
                if os.path.exists(temp_dir):
                    # os.rmdir(temp_dir)
                    import shutil
                    shutil.rmtree(temp_dir)
    else:
        st.markdown('<div class="info-box">📋 Silakan upload file PDF SLIK Report di sidebar</div>', unsafe_allow_html=True)
        
        # Informasi tentang aplikasi
        st.markdown("""
        ### ℹ️ Tentang Aplikasi
        
        Aplikasi ini mengkonversi file PDF SLIK Report menjadi format Excel dengan tahapan:
        
        1. **Read PDF** - Membaca file PDF yang diupload
        2. **Convert to JSON** - Mengkonversi konten PDF ke format JSON
        3. **Read JSON File** - Membaca dan menggabungkan file JSON
        4. **Process Data** - Memproses berbagai jenis data (Kredit, LC, Garansi, Surat, Fasilitas)
        5. **Combine & Clean** - Menggabungkan dan membersihkan data
        6. **Export to Excel** - Menghasilkan file Excel yang diformat
        
        ### 📋 Format Output Excel
        
        File Excel yang dihasilkan akan berisi data dengan kolom-kolom berikut:
        - Nama Group & Debitur
        - BANK & CABANG
        - Kategori Fasilitas
        - Data Plafon, Baki Debet, Suku Bunga
        - Kualitas & Kondisi
        - Data Tanggal (Mulai, Jatuh Tempo, Kondisi, dll)
        - Dan kolom-kolom lainnya sesuai format SLIK Report
        
        ### ⚠️ Catatan Penting
        
        Pastikan file "ZZ Data Kode Bank.xlsx" berada di direktori yang sama dengan aplikasi
        untuk pemrosesan data bank yang optimal.
        """)

if __name__ == "__main__":
    main()
